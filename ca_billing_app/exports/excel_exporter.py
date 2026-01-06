import openpyxl
from openpyxl.utils import get_column_letter

class ExcelExporter:
    def export_to_excel(self, data, headers, filename):
        """
        data: list of dicts or list of lists
        headers: list of column names
        filename: output path
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"
        
        # Write Headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            
        # Write Data
        for row_num, row_data in enumerate(data, 2):
            if isinstance(row_data, dict):
                row_values = [row_data.get(h, '') for h in headers]
            else:
                row_values = row_data # list
                
            for col_num, value in enumerate(row_values, 1):
                ws.cell(row=row_num, column=col_num).value = value
                
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
            
        wb.save(filename)
        return filename
